from datetime import datetime, timedelta
from aiogram import Router, F
from aiogram.filters import Command
from aiogram.types import Message, BufferedInputFile
import pytz
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from database.models import UserRole
from database.queries import AsyncSessionLocal, get_daily_summary, get_monthly_summary, get_monthly_details, get_low_stock_threshold
from keyboards.reply import main_menu
from utils.formatters import fmt_kg, stock_status_text
from config import config

router = Router()
TZ = pytz.timezone(config.TIMEZONE)
REPORT_ROLES = [UserRole.ADMIN, UserRole.MANAGEMENT]

@router.message(Command("report_daily"))
@router.message(F.text == "📊 Daily Report")
async def cmd_report_daily(message: Message, db_user):
    if not (db_user and db_user.is_active and db_user.role in REPORT_ROLES):
        await message.answer("❌ Access denied.")
        return
    local_now = datetime.now(TZ)
    local_date = local_now.date()
    start_local = datetime(local_date.year, local_date.month, local_date.day, 0, 0, 0)
    end_local = start_local + timedelta(days=1)
    start_utc = TZ.localize(start_local).astimezone(pytz.UTC).replace(tzinfo=None)
    end_utc = TZ.localize(end_local).astimezone(pytz.UTC).replace(tzinfo=None)
    async with AsyncSessionLocal() as session:
        summary = await get_daily_summary(session, start_utc, end_utc)
        threshold = await get_low_stock_threshold(session)
    status = stock_status_text(summary["closing_stock"], threshold)
    net = summary["received"] - summary["consumed"] + summary["adjusted_add"] - summary["adjusted_deduct"]
    date_str = local_date.strftime("%d-%b-%Y")
    text = f"📊 <b>DAILY CEMENT REPORT</b>\n━━━━━━━━━━━━━━━━━━━━━━\n📅 Date:           <b>{date_str}</b>\n━━━━━━━━━━━━━━━━━━━━━━\n📥 Received:       <b>{fmt_kg(summary['received'])}</b>\n📤 Consumed:       <b>{fmt_kg(summary['consumed'])}</b>\n📐 m³ Poured:      <b>{summary['cubic_meters']:,.1f} m³</b>\n"
    if summary["adjusted_add"] or summary["adjusted_deduct"]:
        text += f"➕ Adj (Add):      <b>{fmt_kg(summary['adjusted_add'])}</b>\n➖ Adj (Deduct):   <b>{fmt_kg(summary['adjusted_deduct'])}</b>\n"
    text += f"📊 Net Change:     <b>{'+' if net >= 0 else ''}{fmt_kg(net)}</b>\n━━━━━━━━━━━━━━━━━━━━━━\n📦 Closing Stock:  <b>{fmt_kg(summary['closing_stock'])}</b>\n📋 Status:         {status}"
    await message.answer(text, parse_mode="HTML", reply_markup=main_menu(db_user.role))

@router.message(Command("report_monthly"))
@router.message(F.text == "📈 Monthly Report")
async def cmd_report_monthly(message: Message, db_user):
    if not (db_user and db_user.is_active and db_user.role in REPORT_ROLES):
        await message.answer("❌ Access denied.")
        return
    now = datetime.now(TZ)
    year, month = now.year, now.month
    async with AsyncSessionLocal() as session:
        details = await get_monthly_details(session, year, month)
        summary = await get_monthly_summary(session, year, month)
        threshold = await get_low_stock_threshold(session)
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Receipts"
    ws1.append(["Date", "Supplier Name", "Truck Plate", "Volume (kg)"])
    for r in details["receipts"]:
        dt = r.timestamp.replace(tzinfo=pytz.UTC).astimezone(TZ)
        ws1.append([dt.strftime("%Y-%m-%d %H:%M"), r.supplier_name, r.truck_number, r.quantity_kg])
    ws2 = wb.create_sheet("Consumptions")
    ws2.append(["Date", "Project Name", "Concrete Grade", "m³", "Cement Used (kg)", "kg/m³"])
    for c in details["consumptions"]:
        dt = c.timestamp.replace(tzinfo=pytz.UTC).astimezone(TZ)
        ws2.append([dt.strftime("%Y-%m-%d %H:%M"), c.project_name or "", c.concrete_grade or "", c.cubic_meters or 0, c.quantity_kg, c.kg_per_m3 or 0])
    for ws in [ws1, ws2]:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    month_name = datetime(year, month, 1).strftime("%B_%Y")
    status = stock_status_text(summary["closing_stock"], threshold)
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=f"cement_report_{month_name}.xlsx"),
        caption=f"📈 Monthly report for {month_name}\n📥 Received: {fmt_kg(summary['received'])}\n📤 Consumed: {fmt_kg(summary['consumed'])}\n📦 Closing Stock: {fmt_kg(summary['closing_stock'])}\n📋 Status: {status}",
        reply_markup=main_menu(db_user.role)
    )
